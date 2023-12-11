import os
import subprocess

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
import configparser

# Create a configuration object
config = configparser.ConfigParser()

# Read the configuration file
config.read('automation_configurations_xo.cfg')

# Retrieve the values from the 'DEFAULT' section
download_path = config['DEFAULT']['download_path']
from_to_datestrings = config['DEFAULT']['from_to_datestrings']

output_file_path = download_path + 'tasan to zoho xo import ' + from_to_datestrings + '.xlsx'

# Print the retrieved values
print("Download Path:", download_path)
print("From-to Datestrings:", from_to_datestrings)

reloaded_output_file_path = output_file_path
reloaded = pd.read_excel(reloaded_output_file_path, usecols=None, engine="openpyxl", dtype=str)

# Create a new workbook
output_wb = Workbook()
output_ws = output_wb.active

header_row = ['Journal Number Prefix', 'Journal Number Suffix', 'Journal Date', 'Status', 'Journal Type', 'Journal Transaction Type', 'Reference Number', 'Notes', 'Exchange Rate', 'Tax Name', 'Tax Percentage', 'Tax Type', 'Project Name', 'Debit', 'Credit', 'Account', 'Contact Name', 'Service Provider', 'Currency', 'Description']
output_ws.append(header_row)

# Get the column index of column AA
start_column = column_index_from_string('AA')
end_column = column_index_from_string('AT')

# Copy the values from the template (excluding the header) to the new workbook
# Iterate over the rows of the DataFrame without formulas
for row in reloaded.itertuples(): # index=False, name=None): # ???
    values_to_copy = row[start_column:end_column]  # Slice from the start column onwards
    
    output_ws.append(values_to_copy)

# for row in reloaded.iter_rows(values_only=True):
#     # Copy values from column AA and beyond
#     values_to_copy = row[column_index_from_string('AA')-1:]
#     output_ws.append(values_to_copy)

# Save the new workbook with values only
output_values_file_path = download_path + 'tasan to zoho xo import content.xls'
output_wb.save(output_values_file_path)
